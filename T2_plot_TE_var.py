"""
Magnitude T2 fit with alpha comparison
======================================
Model:
    S(TE) = | M0 * alpha * exp(-TE / T2) |

This script:
 - Fits both with alpha free and with alpha fixed to 1
 - Shows both curves on the same plot
 - Saves individual plots and embeds both in Excel
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from tkinter import Tk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# ============================
# 1) Load data
# ============================
def load_txt_file(prompt="Select a TXT file"):
    Tk().withdraw()
    file_path = filedialog.askopenfilename(
        title=prompt,
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
    )
    if not file_path:
        raise FileNotFoundError("❌ No file selected!")
    return file_path

amp_file = load_txt_file("Select TXT file with amplitudes")
te_file = load_txt_file("Select TXT file with echo times (TE)")

def extract_amplitudes_jmrui(path):
    amps = []
    inside = False
    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if line.startswith("Amplitudes"):
                inside = True
                continue
            if inside:
                if line.startswith("Standard deviation of Amplitudes"):
                    break
                for val in line.split():
                    try:
                        amps.append(float(val))
                    except ValueError:
                        pass
    return np.array(amps, dtype=float)

Amplitudes = extract_amplitudes_jmrui(amp_file)

TE = []
with open(te_file, "r") as f:
    for line in f:
        line = line.strip()
        if line:
            try:
                TE.append(float(line))
            except ValueError:
                pass
TE = np.array(TE, dtype=float)

if len(Amplitudes) != len(TE):
    raise ValueError("Mismatch between amplitudes and TE values.")

# ============================
# 2) Model definitions
# ============================
def mag_t2_alpha(TE, M0, T2, alpha):
    return np.abs(M0 * alpha * np.exp(-TE / T2))

def mag_t2_alpha1(TE, M0, T2):
    return np.abs(M0 * np.exp(-TE / T2))

# ============================
# 3) Fit with alpha free
# ============================
M0_guess = np.max(Amplitudes)
T2_guess = np.median(TE)
alpha_guess = 1.0
p0 = [M0_guess, T2_guess, alpha_guess]
lower = [0.0, 1e-6, 0.0]
upper = [M0_guess * 10.0, TE.max() * 100.0, 1.5]

use_fixed_alpha = False
try:
    popt_alpha, pcov_alpha = curve_fit(mag_t2_alpha, TE, Amplitudes,
                                       p0=p0, bounds=(lower, upper), maxfev=20000)
    perr_alpha = np.sqrt(np.abs(np.diag(pcov_alpha)))
    M0_alpha, T2_alpha, alpha_fit = popt_alpha
    M0_err_alpha, T2_err_alpha, alpha_err = perr_alpha
    if not (0 <= alpha_fit <= 1.5):
        raise RuntimeError("alpha out of range, fallback to 1")
except:
    use_fixed_alpha = True
    popt_alpha = [M0_guess, T2_guess, 1.0]
    perr_alpha = [np.nan, np.nan, np.nan]
    M0_alpha, T2_alpha, alpha_fit = popt_alpha
    M0_err_alpha, T2_err_alpha, alpha_err = perr_alpha

# ============================
# 4) Fit with alpha fixed = 1
# ============================
p0_fixed = [M0_guess, T2_guess]
bounds_fixed = ([0.0, 1e-6], [M0_guess * 10.0, TE.max() * 100.0])
popt_fixed, pcov_fixed = curve_fit(mag_t2_alpha1, TE, Amplitudes,
                                   p0=p0_fixed, bounds=bounds_fixed, maxfev=20000)
perr_fixed = np.sqrt(np.abs(np.diag(pcov_fixed)))
M0_fixed, T2_fixed = popt_fixed
M0_err_fixed, T2_err_fixed = perr_fixed

# ============================
# 5) Goodness of fit
# ============================
TE_fit = np.linspace(0, TE.max() * 1.05, 400)
pred_alpha = mag_t2_alpha(TE, M0_alpha, T2_alpha, alpha_fit)
pred_fixed = mag_t2_alpha1(TE, M0_fixed, T2_fixed)

def calc_r2(y, yfit):
    ss_res = np.sum((y - yfit) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    return 1.0 - ss_res / ss_tot

r2_alpha = calc_r2(Amplitudes, pred_alpha)
r2_fixed = calc_r2(Amplitudes, pred_fixed)

# ============================
# 6) Plot both together with T2 ± error
# ============================
plt.figure(figsize=(7, 5))
plt.scatter(TE, Amplitudes, label="Measured data", color="blue")
plt.plot(TE_fit, mag_t2_alpha(TE_fit, M0_alpha, T2_alpha, alpha_fit),
         label=f"Fit (alpha free, alpha={alpha_fit:.2f})", color="red")
plt.plot(TE_fit, mag_t2_alpha1(TE_fit, M0_fixed, T2_fixed),
         label="Fit (alpha=1)", color="green", linestyle="--")
plt.xlabel("Echo Time TE (ms)")
plt.ylabel("Signal Intensity (a.u.)")
plt.title("T2 Magnitude Fit Comparison")
plt.text(0.05, 0.9, f"T2 alpha free = {T2_alpha:.1f} ± {T2_err_alpha:.1f} ms", transform=plt.gca().transAxes, color="red")
plt.text(0.05, 0.83, f"T2 alpha=1   = {T2_fixed:.1f} ± {T2_err_fixed:.1f} ms", transform=plt.gca().transAxes, color="green")
plt.legend()
plt.grid(True, linestyle="--", alpha=0.6)
plt.tight_layout()
plt.show()


# ============================
# 7) Save individual plots
# ============================
Tk().withdraw()
save_choice = messagebox.askyesno("Save Results?", "Do you want to save data, fit results, and plots?")

if save_choice:
    save_path = filedialog.asksaveasfilename(title="Save Excel file", defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if save_path:
        # Save Data + FitResults
        df_data = pd.DataFrame({"TE_ms": TE, "Amplitude": Amplitudes})
        df_fit = pd.DataFrame({
            "Parameter": ["M0_alpha", "T2_alpha", "alpha", "R2_alpha",
                          "M0_fixed", "T2_fixed", "R2_fixed"],
            "Value": [M0_alpha, T2_alpha, alpha_fit, r2_alpha,
                      M0_fixed, T2_fixed, r2_fixed],
            "Error": [M0_err_alpha, T2_err_alpha, alpha_err, "",
                      M0_err_fixed, T2_err_fixed, ""]
        })

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            df_data.to_excel(writer, sheet_name="Data", index=False)
            df_fit.to_excel(writer, sheet_name="FitResults", index=False)

        # Save plots individually
        plot_alpha_file = save_path.replace(".xlsx", "_plot_alpha.png")
        plot_fixed_file = save_path.replace(".xlsx", "_plot_fixed.png")
        plot_combined_file = save_path.replace(".xlsx", "_plot_combined.png")

        # Plot alpha free
        plt.figure(figsize=(7, 5))
        plt.scatter(TE, Amplitudes, label="Data", color="blue")
        plt.plot(TE_fit, mag_t2_alpha(TE_fit, M0_alpha, T2_alpha, alpha_fit), color="red")
        plt.text(0.05, 0.9, f"T2 = {T2_alpha:.1f} ± {T2_err_alpha:.1f} ms", transform=plt.gca().transAxes, color="red")
        plt.title("T2 Fit (alpha free)")
        plt.xlabel("TE (ms)")
        plt.ylabel("Signal")
        plt.tight_layout()
        plt.savefig(plot_alpha_file, dpi=150)
        plt.close()

        # Plot alpha=1
        plt.figure(figsize=(7, 5))
        plt.scatter(TE, Amplitudes, label="Data", color="blue")
        plt.plot(TE_fit, mag_t2_alpha1(TE_fit, M0_fixed, T2_fixed), color="green")
        plt.text(0.05, 0.9, f"T2 = {T2_fixed:.1f} ± {T2_err_fixed:.1f} ms", transform=plt.gca().transAxes,
                 color="green")
        plt.title("T2 Fit (alpha=1)")
        plt.xlabel("TE (ms)")
        plt.ylabel("Signal")
        plt.tight_layout()
        plt.savefig(plot_fixed_file, dpi=150)
        plt.close()

        # Combined plot
        plt.figure(figsize=(7, 5))
        plt.scatter(TE, Amplitudes, label="Data", color="blue")
        plt.plot(TE_fit, mag_t2_alpha(TE_fit, M0_alpha, T2_alpha, alpha_fit),
                 label=f"Fit alpha free (alpha={alpha_fit:.2f})", color="red")
        plt.plot(TE_fit, mag_t2_alpha1(TE_fit, M0_fixed, T2_fixed),
                 label="Fit alpha=1", color="green", linestyle="--")
        plt.text(0.05, 0.9, f"T2 alpha free = {T2_alpha:.1f} ± {T2_err_alpha:.1f} ms", transform=plt.gca().transAxes,
                 color="red")
        plt.text(0.05, 0.83, f"T2 alpha=1   = {T2_fixed:.1f} ± {T2_err_fixed:.1f} ms", transform=plt.gca().transAxes,
                 color="green")
        plt.xlabel("TE (ms)")
        plt.ylabel("Signal Intensity")
        plt.title("T2 Fit Comparison")
        plt.legend()
        plt.grid(True, linestyle="--", alpha=0.6)
        plt.tight_layout()
        plt.savefig(plot_combined_file, dpi=150)
        plt.close()

        # Embed plots into Excel
        wb = load_workbook(save_path)
        ws = wb.create_sheet("Plots")
        ws.add_image(Image(plot_alpha_file), "B2")
        ws.add_image(Image(plot_fixed_file), "B35")
        ws.add_image(Image(plot_combined_file), "B68")
        wb.save(save_path)
        print(f"📁 Results saved with plots to {save_path}")
    else:
        print("❌ Save cancelled.")
else:
    print("ℹ️ Results were not saved.")
