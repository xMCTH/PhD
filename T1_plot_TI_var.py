"""
Signed Inversion-Recovery Fit (with Phase-Aware Amplitudes)
============================================================
Model:
    S(TI) = M0 * (1 - 2 * alpha * exp(-TI / T1))

This script:
 - Reads amplitudes and phases from a jMRUI-style TXT file
 - Converts amplitudes to signed values (phase < 0° → amplitude < 0)
 - Reads TI values from another TXT file (first line = header)
 - Fits M0, T1, alpha (fallback: alpha=1 if free fit fails)
 - Displays and saves three plots:
       1) alpha-free fit
       2) alpha = 1 fit
       3) combined comparison
 - Optionally saves results and plots to Excel
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
from tkinter import Tk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# ============================
# 1) Load data interactively
# ============================
def load_txt_file(prompt):
    Tk().withdraw()
    path = filedialog.askopenfilename(
        title=prompt,
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
    )
    if not path:
        raise FileNotFoundError("❌ No file selected!")
    return path


amp_file = load_txt_file("Select TXT file with amplitudes + phases (jMRUI results)")
ti_file = load_txt_file("Select TXT file with inversion times (first line = header)")

# ============================
# 2) Parse amplitude + phase file
# ============================
def extract_signed_amplitudes_jmrui(path):
    amps, phases = [], []
    inside_amp = inside_phase = False

    with open(path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Amplitudes"):
                inside_amp = True
                continue
            if inside_amp:
                if line.startswith("Standard deviation of Amplitudes"):
                    inside_amp = False
                    continue
                for v in line.split():
                    try:
                        amps.append(float(v))
                    except ValueError:
                        pass

            if line.startswith("Phases"):
                inside_phase = True
                continue
            if inside_phase:
                if line.startswith("Standard deviation of Phases"):
                    inside_phase = False
                    continue
                for v in line.split():
                    try:
                        phases.append(float(v))
                    except ValueError:
                        pass

    if not amps:
        raise ValueError("❌ No amplitudes found.")

    if not phases:
        print("⚠️ No phase data found — using amplitudes as-is.")
        return np.array(amps, float)

    n = min(len(amps), len(phases))
    signed = [a if p >= 0 else -a for a, p in zip(amps[:n], phases[:n])]
    print(f"✅ Extracted {n} signed amplitudes.")
    return np.array(signed, float)


Amplitudes = extract_signed_amplitudes_jmrui(amp_file)

# ============================
# 3) Parse TI file
# ============================
def load_ti_values(path):
    vals = []
    with open(path, "r") as f:
        for line in f.readlines()[1:]:
            try:
                vals.append(float(line.strip()))
            except ValueError:
                pass
    if not vals:
        raise ValueError("❌ No TI values found.")
    print(f"✅ Loaded {len(vals)} TI values.")
    return np.array(vals, float)


TI = load_ti_values(ti_file)

# ============================
# 4) Consistency check
# ============================
if len(TI) != len(Amplitudes):
    raise ValueError("❌ Number of TI values and amplitudes do not match.")

# ============================
# 5) Models
# ============================
def ir_model_signed(TI, M0, T1, alpha):
    return M0 * (1 - 2 * alpha * np.exp(-TI / T1))


def ir_model_alpha1(TI, M0, T1):
    return M0 * (1 - 2 * np.exp(-TI / T1))

# ============================
# 6) Fit
# ============================
M0_guess = np.max(np.abs(Amplitudes))
T1_guess = np.median(TI)

try:
    popt, pcov = curve_fit(
        ir_model_signed, TI, Amplitudes,
        p0=[M0_guess, T1_guess, 1.0],
        bounds=([0, 1e-6, 0], [M0_guess * 10, TI.max() * 100, 1.5]),
        maxfev=20000
    )
    M0_fit, T1_fit, alpha_fit = popt
    M0_err, T1_err, alpha_err = np.sqrt(np.diag(pcov))
except Exception:
    alpha_fit = 1.0
    popt, pcov = curve_fit(
        ir_model_alpha1, TI, Amplitudes,
        p0=[M0_guess, T1_guess],
        bounds=([0, 1e-6], [M0_guess * 10, TI.max() * 100]),
        maxfev=20000
    )
    M0_fit, T1_fit = popt
    M0_err, T1_err = np.sqrt(np.diag(pcov))
    alpha_err = np.nan

# α = 1 comparison fit
popt_fixed, pcov_fixed = curve_fit(
    ir_model_alpha1, TI, Amplitudes,
    p0=[M0_guess, T1_guess],
    bounds=([0, 1e-6], [M0_guess * 10, TI.max() * 100]),
    maxfev=20000
)
M0_fixed, T1_fixed = popt_fixed
M0_err_fixed, T1_err_fixed = np.sqrt(np.diag(pcov_fixed))

# ============================
# 7) Plot & store three figures
# ============================
TI_fit = np.linspace(0, TI.max() * 1.05, 400)
fit_alpha = ir_model_signed(TI_fit, M0_fit, T1_fit, alpha_fit)
fit_fixed = ir_model_alpha1(TI_fit, M0_fixed, T1_fixed)

figures = {}

# --- α-free
fig, ax = plt.subplots(figsize=(7, 5))
ax.scatter(TI, Amplitudes, label="Data")
ax.plot(TI_fit, fit_alpha, label=f"α-free (α={alpha_fit:.2f})")
ax.set_title("T1 Inversion Recovery – α-free")
ax.set_xlabel("TI (ms)")
ax.set_ylabel("Signal (a.u.)")
ax.text(0.05, 0.9, f"T1 = {T1_fit:.1f} ± {T1_err:.1f} ms", transform=ax.transAxes)
ax.legend(); ax.grid(True, linestyle="--", alpha=0.6)
fig.tight_layout()
plt.show()
figures["alpha_free"] = fig

# --- α = 1
fig, ax = plt.subplots(figsize=(7, 5))
ax.scatter(TI, Amplitudes, label="Data")
ax.plot(TI_fit, fit_fixed, "--", label="α = 1")
ax.set_title("T1 Inversion Recovery – α = 1")
ax.set_xlabel("TI (ms)")
ax.set_ylabel("Signal (a.u.)")
ax.text(0.05, 0.9, f"T1 = {T1_fixed:.1f} ± {T1_err_fixed:.1f} ms", transform=ax.transAxes)
ax.legend(); ax.grid(True, linestyle="--", alpha=0.6)
fig.tight_layout()
plt.show()
figures["alpha_1"] = fig

# --- Combined
fig, ax = plt.subplots(figsize=(7, 5))
ax.scatter(TI, Amplitudes, label="Data")
ax.plot(TI_fit, fit_alpha, label=f"α-free (α={alpha_fit:.2f})")
ax.plot(TI_fit, fit_fixed, "--", label="α = 1")
ax.set_title("T1 Inversion Recovery – Combined")
ax.set_xlabel("TI (ms)")
ax.set_ylabel("Signal (a.u.)")
ax.text(
    0.05, 0.88,
    f"T1 α-free = {T1_fit:.1f} ± {T1_err:.1f} ms\n"
    f"T1 α=1    = {T1_fixed:.1f} ± {T1_err_fixed:.1f} ms",
    transform=ax.transAxes
)
ax.legend(); ax.grid(True, linestyle="--", alpha=0.6)
fig.tight_layout()
plt.show()
figures["combined"] = fig

# ============================
# 8) Optional Excel export
# ============================
Tk().withdraw()
if messagebox.askyesno("Save Results?", "Save data, fit results, and plots into Excel?"):

    save_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if save_path:
        df_data = pd.DataFrame({"TI_ms": TI, "Amplitude": Amplitudes})
        df_fit = pd.DataFrame({
            "Parameter": ["M0", "T1", "alpha"],
            "Value": [M0_fit, T1_fit, alpha_fit],
            "Error": [M0_err, T1_err, alpha_err]
        })
        df_fit_fixed = pd.DataFrame({
            "Parameter": ["M0_fixed", "T1_fixed"],
            "Value": [M0_fixed, T1_fixed],
            "Error": [M0_err_fixed, T1_err_fixed]
        })

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            df_data.to_excel(writer, sheet_name="Data", index=False)
            df_fit.to_excel(writer, sheet_name="Fit_alpha_free", index=False)
            df_fit_fixed.to_excel(writer, sheet_name="Fit_alpha_1", index=False)

        wb = load_workbook(save_path)
        ws = wb.create_sheet("Plots")

        row = 1
        for name, fig in figures.items():
            img_path = save_path.replace(".xlsx", f"_{name}.png")
            fig.savefig(img_path, dpi=150)
            ws.add_image(Image(img_path), f"A{row}")
            row += 25
            plt.close(fig)

        wb.save(save_path)
        print(f"📁 Results and plots saved to {save_path}")
