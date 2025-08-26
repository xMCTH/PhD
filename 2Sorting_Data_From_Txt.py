"""
Phantom vs. Subject Measurement Processing Script
-------------------------------------------------

This script processes two fitted MRSI measurement `.txt` files: one from a phantom and one from a subject.
It extracts voxel-wise metabolite data, allows the user to filter metabolites of interest,
generates quick previews and optionally exports results to Excel with profile corrected results.

Workflow:
1. Prompts the user to select:
   - Phantom measurement TXT file
   - Subject measurement TXT file
2. Parses the files into structured entries, extracting voxel coordinates (x, y, z)
   and measurement parameters.
3. Lists all available metabolites, then asks the user to select one or more metabolites
   for analysis.
4. Displays a preview of filtered data (coordinate, metabolite, area, and LDamping).
5. If the user opts to save to Excel:
   - Creates a workbook containing:
     - An **"all" sheet** with phantom (_p) and subject (_s) data aligned by voxel coordinates.
       Includes formulas for:
         • Height = Area / LDamping
         • FWHM = LDamping / π
         • I0ps = Height * EXP(TE / FWHM)
         • I0 = I0ps * (1 - EXP(-TR / T1))
     - Per-slice sheets (z-value separated) for both phantom and subject data.
     - Extra calculated concentration columns (`Pc [mM]`, `c [mM]`).
   - Ensures all formulas dynamically link across sheets for consistency.
   - Saves the workbook to a user-selected `.xlsx` file.

Dependencies:
- Python standard library (`os`, `tkinter`)
- Third-party libraries:
  • numpy
  • pandas
  • matplotlib
  • openpyxl

Usage:
- Run the script in Python.
- Select two `.txt` measurement files when prompted.
- Choose metabolites of interest.
- Optionally export results to Excel with formulas grouped by z-value.


"""


import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from tkinter import filedialog, Tk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# --- FILE SELECTION ---
root = Tk()
root.withdraw()  # Hide the main window

# Select phantom measurement file
file_path1 = filedialog.askopenfilename(
    title="Select Phantom Measurement TXT File",
    filetypes=[("Text Files", "*.txt")]
)
if not file_path1:
    print("No phantom measurement file selected.")
    exit()

# Select subject measurement file
file_path2 = filedialog.askopenfilename(
    title="Select Subject Measurement TXT File",
    filetypes=[("Text Files", "*.txt")]
)
if not file_path2:
    print("No subject measurement file selected.")
    exit()

# --- VALIDATE FILES ---
if not (os.path.exists(file_path1) and os.path.exists(file_path2)):
    print("One or both .txt file paths are invalid.")
    exit()

print(f"\nUsing files:\n1. {file_path1}\n2. {file_path2}")


if not (os.path.exists(file_path1) and os.path.exists(file_path2)):
    print("One or both .txt file paths are invalid.")
    exit()

print(f"\nUsing files:\n1. {file_path1}\n2. {file_path2}")

def parse_txt_file(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    blocks = content.strip().split('###')
    entries = []
    local_headers = None
    for block in blocks:
        lines = block.strip().split('\n')
        if not lines or not lines[0].startswith("Coord"):
            continue
        local_headers = lines[0].split('\t')
        for line in lines[1:]:
            parts = line.split('\t')
            if len(parts) >= len(local_headers):
                entry = dict(zip(local_headers, parts))
                try:
                    x, y, z = map(int, entry['Coord'].split('_'))
                    entry.update({'x': x, 'y': y, 'z': z})
                    entries.append(entry)
                except Exception:
                    continue
    return entries, local_headers

# Parse both files
data_entries1, headers = parse_txt_file(file_path1)
data_entries2, _ = parse_txt_file(file_path2)

sorted_data1 = sorted(data_entries1, key=lambda e: (e['z'], e['x'], e['y']))
sorted_data2 = sorted(data_entries2, key=lambda e: (e['z'], e['x'], e['y']))

metabolites = sorted({e['Metabolite'] for e in sorted_data1})
print("\nAvailable Metabolites:")
for m in metabolites:
    print(f" - {m}")
selected = input("\nType metabolite names (comma-separated), e.g. Main: ")
sel_set = {m.strip() for m in selected.split(',') if m.strip()}

filtered1 = [e for e in sorted_data1 if e['Metabolite'] in sel_set]
filtered2 = [e for e in sorted_data2 if e['Metabolite'] in sel_set]
if not filtered1 or not filtered2:
    print("No matching data in one or both files. Exiting.")
    exit()

_df = pd.DataFrame(filtered1)
print("\n\U0001F9BE Preview:")
print(_df[['Coord', 'Metabolite', 'Area', 'LDamping']].head(10))

main = [d for d in filtered1 if d['Metabolite'] == 'Main']
if main:
    dfm = pd.DataFrame(main)
    dfm['Area'] = pd.to_numeric(dfm['Area'], errors='coerce')
    pt = dfm.pivot_table(index='y', columns='x', values='Area', aggfunc='first')
    plt.figure(figsize=(10, 8))
    plt.imshow(pt, origin='lower', aspect='auto', cmap='viridis')
    plt.colorbar(label='Area')
    plt.title("Heatmap of 'Main' Area")
    plt.tight_layout()
    plt.show()

if input("\nSave as Excel with formulas by z-value? (y/n): ").strip().lower() == 'y':
    root = Tk()
    root.withdraw()
    excel_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])
    if excel_path:
        wb = Workbook()
        wb.remove(wb.active)

        header_cols = headers + ['Height', 'FWHM', 'I0ps', 'I0']
        area_idx = header_cols.index('Area')
        ld_idx = header_cols.index('LDamping')
        h_idx = header_cols.index('Height')
        f_idx = header_cols.index('FWHM')
        i0_idx = header_cols.index('I0')

        def col_letter(idx): return get_column_letter(idx + 1)

        df1 = pd.DataFrame(filtered1)
        df2 = pd.DataFrame(filtered2)

        def extract_coord_key(coord_str):
            try:
                x, y, z = map(int, coord_str.split('_'))
                return (z, x, y)
            except:
                return (9999, 9999, 9999)

        all_coords = sorted(set(df1['Coord']).union(df2['Coord']), key=extract_coord_key)
        coord_to_entries1 = {e['Coord']: e for e in filtered1}
        coord_to_entries2 = {e['Coord']: e for e in filtered2}

        ws_all = wb.create_sheet(title="all")
        ws_all.cell(row=1, column=1, value="TRp")
        ws_all.cell(row=2, column=1, value="T1p")
        ws_all.cell(row=3, column=1, value="TEp")
        ws_all.cell(row=1, column=3, value="TRs")
        ws_all.cell(row=2, column=3, value="T1s")
        ws_all.cell(row=3, column=3, value="TEs")

        for c, header in enumerate(header_cols, start=1):
            ws_all.cell(row=4, column=c, value=f"{header}_p")
        for c, header in enumerate(header_cols, start=1):
            ws_all.cell(row=4, column=c + 18, value=f"{header}_s")

        row_pointer = 5
        for coord in all_coords:
            r1 = coord_to_entries1.get(coord, None)
            r2 = coord_to_entries2.get(coord, None)

            if r1:
                for c, col in enumerate(headers, start=1):
                    ws_all.cell(row=row_pointer, column=c, value=r1.get(col, ''))
                a_col = col_letter(area_idx)
                l_col = col_letter(ld_idx)
                h_col = col_letter(h_idx)
                f_col = col_letter(f_idx)
                ws_all.cell(row=row_pointer, column=h_idx + 1, value=f"={a_col}{row_pointer}/{l_col}{row_pointer}")
                ws_all.cell(row=row_pointer, column=f_idx + 1, value=f"={l_col}{row_pointer}/PI()")
                ws_all.cell(row=row_pointer, column=f_idx + 2, value=f"={h_col}{row_pointer}*EXP($B$3/{f_col}{row_pointer})")
                i0_col = f_idx + 3
                left_col_letter = get_column_letter(i0_col - 1)
                ws_all.cell(row=row_pointer, column=i0_col,
                            value=f"={left_col_letter}{row_pointer}*(1-EXP(-$B$1/$B$2))")

            if r2:
                offset = 18
                for c, col in enumerate(headers, start=1):
                    ws_all.cell(row=row_pointer, column=c + offset, value=r2.get(col, ''))
                ws_all.cell(row=row_pointer, column=h_idx + 1 + offset, value=f"=V{row_pointer}/Y{row_pointer}")
                ws_all.cell(row=row_pointer, column=f_idx + 1 + offset, value=f"=Y{row_pointer}/PI()")
                ws_all.cell(row=row_pointer, column=f_idx + 2 + offset, value=f"=AE{row_pointer}*EXP($D$3/AF{row_pointer})")
                ws_all.cell(row=row_pointer, column=f_idx + 3 + offset,
                            value=f"=AG{row_pointer}*(1-EXP(-$D$1/$D$2))")

            row_pointer += 1


        def write_z_sheet(wb, sheet_name, df_z, is_primary):
            ws = wb.create_sheet(title=sheet_name)

            # Copy parameter headers from 'all' sheet
            ws['A1'] = "TRp"
            ws['A2'] = "T1p"
            ws['A3'] = "TEp"
            ws['C1'] = "TRs"
            ws['C2'] = "T1s"
            ws['C3'] = "TEs"

            # Link values from 'all' sheet
            ws['B1'] = "=all!B1"
            ws['B2'] = "=all!B2"
            ws['B3'] = "=all!B3"
            ws['D1'] = "=all!D1"
            ws['D2'] = "=all!D2"
            ws['D3'] = "=all!D3"
            ws['F1'] = "=all!F1"

            for c, header in enumerate(header_cols, start=1):
                ws.cell(row=4, column=c, value=header)

            for r, entry in enumerate(df_z.itertuples(index=False), start=5):
                for c, col in enumerate(headers, start=1):
                    ws.cell(row=r, column=c, value=getattr(entry, col, ''))

                a_col = col_letter(area_idx)
                l_col = col_letter(ld_idx)
                h_col = col_letter(h_idx)
                f_col = col_letter(f_idx)

                ws.cell(row=r, column=h_idx + 1, value=f"={a_col}{r}/{l_col}{r}")
                ws.cell(row=r, column=f_idx + 1, value=f"={l_col}{r}/PI()")
                ws.cell(row=r, column=f_idx + 2, value=f"={h_col}{r}*EXP($B$3/{f_col}{r})")

                i0_col = f_idx + 3
                ws.cell(row=r, column=i0_col, value=f"=N{r}*(1-EXP(-$B$1/$B$2))")


        for z_val in sorted(df1['z'].unique()):
            df_z = df1[df1['z'] == z_val]
            write_z_sheet(wb, f"z{z_val}p", df_z, is_primary=True)

        for z_val in sorted(df2['z'].unique()):
            df_z = df2[df2['z'] == z_val]
            write_z_sheet(wb, f"z{z_val}s", df_z, is_primary=False)

        for ws in wb.worksheets:
            ws.cell(row=1, column=5, value="Pc [mM]")
            if ws.title != "all":
                ws.cell(row=1, column=6, value="='all'!F1")

        for ws in wb.worksheets:
            ws.insert_cols(i0_idx + 2)
            ws.cell(row=4, column=i0_idx + 2, value="c [mM]")
            last_row = ws.max_row
            if ws.title == "all":
                for r in range(5, last_row + 1):
                    formula = f'=IF(AND(O{r}<>0, AH{r}<>0), (AH{r}/O{r})*$F$1, "")'
                    ws.cell(row=r, column=i0_idx + 2, value=formula)
            else:
                for r in range(5, last_row + 1):
                    ws.cell(row=r, column=i0_idx + 2, value="")



        wb.save(excel_path)
        print(f"\nSaved Excel file to {excel_path}")
    else:
        print("Save cancelled.")
else:
    print("Exiting without saving.")
